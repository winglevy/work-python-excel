# -*- coding: utf-8 -*-
"""
Python 操作 Excel 经典案例
功能包含：
1. 使用 openpyxl 创建 Excel 文件并写入数据
2. 读取 Excel 数据
3. 修改现有 Excel 文件
4. 使用 pandas 进行数据分析
"""

# 安装依赖库（如果未安装）
# pip install openpyxl pandas

import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

# ====================
# 1. 使用 openpyxl 创建 Excel 文件
# ====================
def create_excel_file():
    # 创建 Workbook 和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生成绩"

    # 写入表头
    headers = ["姓名", "数学", "语文", "英语"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # 写入示例数据
    students = [
        ("张三", 85, 90, 88),
        ("李四", 92, 85, 95),
        ("王五", 78, 82, 80)
    ]
    for row_num, student in enumerate(students, 2):
        for col_num, value in enumerate(student, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 保存文件
    wb.save("student_scores.xlsx")
    print("Excel 文件创建成功！")

# ====================
# 2. 读取 Excel 数据
# ====================
def read_excel_file():
    wb = openpyxl.load_workbook("student_scores.xlsx")
    ws = wb.active

    # 读取所有数据
    data = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        data.append(row)
    
    print("\n读取到的数据：")
    for row in data:
        print(row)

    return data

# ====================
# 3. 修改 Excel 文件
# ====================
def modify_excel_file():
    wb = openpyxl.load_workbook("student_scores.xlsx")
    ws = wb.active

    # 添加总分列
    ws["E1"] = "总分"
    for row in range(2, ws.max_row + 1):
        total = sum(ws.cell(row=row, column=col).value for col in range(2, 5))
        ws.cell(row=row, column=5, value=total)
        # 使用公式的写法：ws.cell(row=row, column=5, value=f"=SUM(B{row}:D{row})")

    # 保存为新文件
    wb.save("student_scores_modified.xlsx")
    print("\n文件修改完成，已保存新版本！")

# ====================
# 4. 使用 pandas 分析数据
# ====================
def analyze_with_pandas():
    # 读取 Excel 文件
    df = pd.read_excel("student_scores_modified.xlsx")
    
    # 添加平均分列
    df["平均分"] = df[["数学", "语文", "英语"]].mean(axis=1).round(2)
    
    # 数据分析
    print("\n数据分析结果：")
    print(f"数学平均分：{df['数学'].mean():.2f}")
    print(f"语文最高分：{df['语文'].max()}")
    print(f"英语最低分：{df['英语'].min()}")
    
    # 保存分析结果
    df.to_excel("student_analysis.xlsx", index=False)
    print("分析结果已保存！")

# ====================
# 主程序
# ====================
if __name__ == "__main__":
    create_excel_file()     # 创建示例文件
    read_excel_file()       # 读取数据演示
    modify_excel_file()     # 修改文件演示
    analyze_with_pandas()   # 使用 pandas 分析
